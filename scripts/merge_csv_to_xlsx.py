#!/usr/bin/env python3
"""Merge CSV files into a single XLSX workbook.

- One worksheet per CSV file
- Optional recursive file discovery
- Safe worksheet naming (Excel constraints)
- Includes a manifest worksheet mapping sheet names to source files

Example:
    python resubmission/scripts/merge_csv_to_xlsx.py \
        --input-dir resubmission/results/plots \
        --output-xlsx resubmission/results/plots/statistics_tables_unified.xlsx
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Iterable, List, Tuple

from openpyxl import Workbook


def discover_csv_files(input_dir: Path, recursive: bool) -> List[Path]:
    pattern = "**/*.csv" if recursive else "*.csv"
    return sorted(p for p in input_dir.glob(pattern) if p.is_file())


def make_unique_sheet_name(raw_name: str, used: set[str]) -> str:
    invalid = set('[]:*?/\\')
    cleaned = "".join("_" if ch in invalid else ch for ch in raw_name).strip()
    if not cleaned:
        cleaned = "Sheet"

    base = cleaned[:31]
    name = base
    idx = 1
    while name in used:
        suffix = f"_{idx}"
        name = (base[: 31 - len(suffix)] + suffix) if len(base) + len(suffix) > 31 else (base + suffix)
        idx += 1

    used.add(name)
    return name


def read_csv_rows(csv_path: Path) -> Tuple[List[str], List[List[str]]]:
    # Attempt dialect sniffing first, fallback to comma if sniff fails.
    with csv_path.open("r", newline="", encoding="utf-8") as fh:
        sample = fh.read(4096)
        fh.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel

        reader = csv.reader(fh, dialect)
        rows = list(reader)

    if not rows:
        return [], []

    header = rows[0]
    data_rows = rows[1:] if len(rows) > 1 else []
    return header, data_rows


def write_workbook(csv_files: Iterable[Path], input_dir: Path, output_xlsx: Path) -> int:
    wb = Workbook()

    # Remove default sheet; create content sheets explicitly.
    default_ws = wb.active
    wb.remove(default_ws)

    used_sheet_names: set[str] = set()
    manifest_rows: List[Tuple[str, str, int, int]] = []

    for csv_file in csv_files:
        relative_path = csv_file.relative_to(input_dir)
        sheet_name = make_unique_sheet_name(relative_path.stem, used_sheet_names)

        header, data_rows = read_csv_rows(csv_file)
        ws = wb.create_sheet(title=sheet_name)

        if header:
            ws.append(header)
        for row in data_rows:
            ws.append(row)

        n_rows = len(data_rows)
        n_cols = len(header) if header else 0
        manifest_rows.append((sheet_name, str(relative_path), n_rows, n_cols))

    # Add manifest as last sheet.
    manifest = wb.create_sheet(title="manifest")
    manifest.append(["sheet_name", "csv_file", "rows", "columns"])
    for row in manifest_rows:
        manifest.append(list(row))

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)
    return len(manifest_rows)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Merge CSV files into one XLSX workbook.")
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=Path("resubmission/results/plots"),
        help="Directory containing CSV files (default: resubmission/results/plots).",
    )
    parser.add_argument(
        "--output-xlsx",
        type=Path,
        default=Path("resubmission/results/plots/statistics_tables_unified.xlsx"),
        help="Output XLSX path.",
    )
    parser.add_argument(
        "--no-recursive",
        action="store_true",
        help="Only include CSVs directly under input-dir (skip subdirectories).",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_dir: Path = args.input_dir
    output_xlsx: Path = args.output_xlsx
    recursive = not args.no_recursive

    if not input_dir.exists() or not input_dir.is_dir():
        raise SystemExit(f"Input directory not found or not a directory: {input_dir}")

    csv_files = discover_csv_files(input_dir=input_dir, recursive=recursive)
    if not csv_files:
        raise SystemExit(f"No CSV files found in: {input_dir}")

    sheet_count = write_workbook(csv_files=csv_files, input_dir=input_dir, output_xlsx=output_xlsx)
    print(f"Created workbook: {output_xlsx}")
    print(f"CSV files merged: {sheet_count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
